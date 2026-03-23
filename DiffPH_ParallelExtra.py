import os
import time
import cv2
import numpy as np
import matplotlib.pyplot as plt
import concurrent.futures
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =========================================================
# SETTINGS
# =========================================================
# Edit these if needed
folder_ph12 = r"C:\Users\aprko\Desktop\Group_Project\PH 12 Images"
folder_ph13 = r"C:\Users\aprko\Desktop\Group_Project\PH 13 Images"

# If both sets are actually in the same folder, just make both paths the same.
# Example:
# folder_ph12 = r"C:\Users\aprko\Desktop\Group_Project\PH 12 Images"
# folder_ph13 = r"C:\Users\aprko\Desktop\Group_Project\PH 12 Images"

datasets = {
    "pH 12": {
        "folder": folder_ph12,
        "start_index": 1,
        "end_index": 16
    },
    "pH 13": {
        "folder": folder_ph13,
        "start_index": 18,
        "end_index": 32
    }
}

n_runs = 10
max_workers = None  # None lets ThreadPoolExecutor choose automatically

# HSV thresholds
lower_blue = np.array([95, 50, 50])
upper_blue = np.array([145, 255, 255])

# Excel output
excel_path = os.path.join(os.path.dirname(folder_ph12), "parallel_processing_timing_summary.xlsx")

# =========================================================
# HELPER FUNCTIONS
# =========================================================
def find_image_path(folder, i):
    """
    Try both 5-digit and 4-digit APC filename styles.
    Returns the first existing path, otherwise None.
    """
    candidates = [
        os.path.join(folder, f"APC_{i:05d}.jpg"),
        os.path.join(folder, f"APC_{i:04d}.jpg")
    ]

    for path in candidates:
        if os.path.exists(path):
            return path

    return None


def get_output_folders(base_folder, dataset_label):
    """
    Create separate output folders for sequential and parallel results.
    """
    safe_label = dataset_label.replace(" ", "_")
    seq_folder = os.path.join(base_folder, f"{safe_label}_Sequential_Output")
    par_folder = os.path.join(base_folder, f"{safe_label}_Parallel_Output")

    os.makedirs(seq_folder, exist_ok=True)
    os.makedirs(par_folder, exist_ok=True)

    return seq_folder, par_folder


def process_image(folder, i, output_folder):
    """
    Read image, isolate blue regions, reduce noise, save output.
    """
    path = find_image_path(folder, i)

    if path is None:
        return i, False, "not found"

    img = cv2.imread(path)

    if img is None:
        return i, False, "read failed"

    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)

    mask = cv2.inRange(hsv, lower_blue, upper_blue)

    kernel = np.ones((3, 3), np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
    mask = cv2.GaussianBlur(mask, (5, 5), 0)

    result = cv2.bitwise_and(img, img, mask=mask)

    base_name = os.path.splitext(os.path.basename(path))[0]
    output_name = f"{base_name}_NoiseReduced.jpg"
    output_path = os.path.join(output_folder, output_name)

    cv2.imwrite(output_path, result)

    return i, True, output_name


def run_sequential(folder, start_index, end_index, output_folder):
    """
    Process images one-by-one and return elapsed time.
    """
    t0 = time.perf_counter()

    for i in range(start_index, end_index + 1):
        process_image(folder, i, output_folder)

    t1 = time.perf_counter()
    return t1 - t0


def run_parallel(folder, start_index, end_index, output_folder):
    """
    Process images in parallel and return elapsed time.
    """
    t0 = time.perf_counter()

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [
            executor.submit(process_image, folder, i, output_folder)
            for i in range(start_index, end_index + 1)
        ]

        for future in concurrent.futures.as_completed(futures):
            future.result()

    t1 = time.perf_counter()
    return t1 - t0


def format_excel_sheet(ws):
    """
    Apply simple professional formatting to a worksheet.
    """
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="808080")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = max_len + 4


def create_excel_workbook(all_results, save_path):
    """
    Save raw timings and summary into one formatted Excel workbook.
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Raw sheets
    for dataset_label, result in all_results.items():
        sheet_name = dataset_label.replace(" ", "") + "_raw"
        ws = wb.create_sheet(title=sheet_name)

        ws.append(["Run", "Sequential Time (s)", "Parallel Time (s)", "Speedup (Seq/Par)", "Improvement (%)"])

        seq_times = result["sequential_times"]
        par_times = result["parallel_times"]

        for run_idx, (seq_t, par_t) in enumerate(zip(seq_times, par_times), start=1):
            speedup = seq_t / par_t if par_t > 0 else None
            improvement = ((seq_t - par_t) / seq_t) * 100 if seq_t > 0 else None
            ws.append([run_idx, seq_t, par_t, speedup, improvement])

        format_excel_sheet(ws)

    # Summary sheet
    ws = wb.create_sheet(title="Summary")
    ws.append([
        "Dataset",
        "Sequential Mean (s)",
        "Sequential SD (s)",
        "Parallel Mean (s)",
        "Parallel SD (s)",
        "Mean Speedup (Seq/Par)",
        "Mean Improvement (%)"
    ])

    for dataset_label, result in all_results.items():
        ws.append([
            dataset_label,
            result["seq_mean"],
            result["seq_std"],
            result["par_mean"],
            result["par_std"],
            result["speedup_mean"],
            result["improvement_mean"]
        ])

    format_excel_sheet(ws)
    wb.save(save_path)


def make_summary_table_plot(all_results):
    """
    Create a report-style matplotlib table.
    """
    rows = []
    for dataset_label, result in all_results.items():
        rows.append([
            dataset_label,
            f'{result["seq_mean"]:.4f} ± {result["seq_std"]:.4f}',
            f'{result["par_mean"]:.4f} ± {result["par_std"]:.4f}',
            f'{result["speedup_mean"]:.2f}×',
            f'{result["improvement_mean"]:.1f}%'
        ])

    fig, ax = plt.subplots(figsize=(11, 3.8))
    ax.axis("off")

    table = ax.table(
        cellText=rows,
        colLabels=[
            "Dataset",
            "Sequential Runtime, mean ± SD (s)",
            "Parallel Runtime, mean ± SD (s)",
            "Mean Speedup",
            "Mean Improvement"
        ],
        cellLoc="center",
        loc="center"
    )

    table.auto_set_font_size(False)
    table.set_fontsize(10.5)
    table.scale(1.15, 1.8)

    for (r, c), cell in table.get_celld().items():
        if r == 0:
            cell.set_text_props(weight="bold")
            cell.set_facecolor("#D9EAF7")
        cell.set_edgecolor("gray")

    plt.title("Runtime Summary for Sequential and Parallel Image Processing", fontsize=14, pad=16)
    plt.tight_layout()
    plt.show()


def make_grouped_bar_plot(all_results):
    """
    Create grouped bar chart with error bars and exact mean values.
    """
    dataset_names = list(all_results.keys())

    seq_means = [all_results[name]["seq_mean"] for name in dataset_names]
    par_means = [all_results[name]["par_mean"] for name in dataset_names]

    seq_stds = [all_results[name]["seq_std"] for name in dataset_names]
    par_stds = [all_results[name]["par_std"] for name in dataset_names]

    x = np.arange(len(dataset_names))
    width = 0.34

    fig, ax = plt.subplots(figsize=(10, 6))

    bars1 = ax.bar(
        x - width / 2,
        seq_means,
        width,
        yerr=seq_stds,
        capsize=7,
        label="Sequential"
    )

    bars2 = ax.bar(
        x + width / 2,
        par_means,
        width,
        yerr=par_stds,
        capsize=7,
        label="Parallel"
    )

    ax.set_title("Comparison of Mean Runtime for Sequential and Parallel Processing", fontsize=14)
    ax.set_xlabel("Dataset", fontsize=12)
    ax.set_ylabel("Mean Execution Time (seconds)", fontsize=12)
    ax.set_xticks(x)
    ax.set_xticklabels(dataset_names, fontsize=11)
    ax.legend()
    ax.grid(axis="y", linestyle="--", alpha=0.6)

    # Write exact values above bars
    for bars, stds in [(bars1, seq_stds), (bars2, par_stds)]:
        for bar, std in zip(bars, stds):
            height = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                height + std + 0.002,
                f"{height:.4f} s",
                ha="center",
                va="bottom",
                fontsize=10
            )

    # Add per-dataset speedup text
    summary_lines = []
    for dataset_label in dataset_names:
        speedup = all_results[dataset_label]["speedup_mean"]
        improvement = all_results[dataset_label]["improvement_mean"]
        summary_lines.append(f"{dataset_label}: {speedup:.2f}×, {improvement:.1f}% faster")

    ax.text(
        0.98, 0.97,
        "\n".join(summary_lines),
        transform=ax.transAxes,
        ha="right",
        va="top",
        fontsize=10.5,
        bbox=dict(boxstyle="round", facecolor="white", alpha=0.9)
    )

    plt.tight_layout()
    plt.show()


def make_per_run_plot(all_results):
    """
    Plot runtime for each repeated run for both datasets.
    """
    fig, axes = plt.subplots(1, 2, figsize=(13, 5), sharey=True)

    for ax, (dataset_label, result) in zip(axes, all_results.items()):
        runs = np.arange(1, len(result["sequential_times"]) + 1)

        ax.plot(runs, result["sequential_times"], marker="o", label="Sequential")
        ax.plot(runs, result["parallel_times"], marker="o", label="Parallel")

        ax.set_title(f"Repeated Runtime Measurements: {dataset_label}", fontsize=12)
        ax.set_xlabel("Run Number", fontsize=11)
        ax.set_ylabel("Execution Time (seconds)", fontsize=11)
        ax.grid(True, linestyle="--", alpha=0.6)
        ax.legend()

    plt.tight_layout()
    plt.show()


# =========================================================
# MAIN
# =========================================================
if __name__ == "__main__":
    all_results = {}

    for dataset_label, info in datasets.items():
        folder = info["folder"]
        start_index = info["start_index"]
        end_index = info["end_index"]

        print(f"\nProcessing {dataset_label}")
        print(f"Folder: {folder}")
        print(f"Image range: {start_index} to {end_index}")

        seq_folder, par_folder = get_output_folders(folder, dataset_label)

        sequential_times = []
        parallel_times = []

        for run in range(1, n_runs + 1):
            seq_time = run_sequential(folder, start_index, end_index, seq_folder)
            par_time = run_parallel(folder, start_index, end_index, par_folder)

            sequential_times.append(seq_time)
            parallel_times.append(par_time)

            print(
                f"Run {run:02d} | "
                f"Sequential = {seq_time:.6f} s | "
                f"Parallel = {par_time:.6f} s"
            )

        seq_mean = float(np.mean(sequential_times))
        seq_std = float(np.std(sequential_times, ddof=1)) if len(sequential_times) > 1 else 0.0

        par_mean = float(np.mean(parallel_times))
        par_std = float(np.std(parallel_times, ddof=1)) if len(parallel_times) > 1 else 0.0

        speedup_mean = seq_mean / par_mean if par_mean > 0 else np.inf
        improvement_mean = ((seq_mean - par_mean) / seq_mean) * 100 if seq_mean > 0 else 0.0

        all_results[dataset_label] = {
            "sequential_times": sequential_times,
            "parallel_times": parallel_times,
            "seq_mean": seq_mean,
            "seq_std": seq_std,
            "par_mean": par_mean,
            "par_std": par_std,
            "speedup_mean": speedup_mean,
            "improvement_mean": improvement_mean
        }

    # Print summary
    print("\n================ FINAL SUMMARY ================")
    for dataset_label, result in all_results.items():
        print(f"\n{dataset_label}")
        print(f"Sequential mean ± SD: {result['seq_mean']:.6f} ± {result['seq_std']:.6f} s")
        print(f"Parallel mean ± SD:   {result['par_mean']:.6f} ± {result['par_std']:.6f} s")
        print(f"Mean speedup:         {result['speedup_mean']:.2f}x")
        print(f"Mean improvement:     {result['improvement_mean']:.2f}%")

    # Save Excel workbook
    create_excel_workbook(all_results, excel_path)
    print(f"\nExcel workbook saved to:\n{excel_path}")

    # Plots
    make_summary_table_plot(all_results)
    make_grouped_bar_plot(all_results)
    make_per_run_plot(all_results)